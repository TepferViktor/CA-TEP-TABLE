import requests
import openpyxl
from openpyxl import load_workbook
import json
def parse_SHIB():
    headers = {
        "Accept": "*/*",
        "Accept-Encoding": "gzip, deflate, br",
        "Accept-Language": "en-GB,en-US;q=0.9,en;q=0.8",
        "Cache-Control": "no-cache",
        "Connection": "keep-alive",
        "Content-Length": "123",
        "content-type": "application/json",
        "Host": "p2p.binance.com",
        "Origin": "https://p2p.binance.com",
        "Pragma": "no-cache",
        "TE": "Trailers",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:88.0) Gecko/20100101 Firefox/88.0"
    }

    data = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'TinkoffNew',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'SHIB',
    'fiat': 'RUB',
    'tradeType': 'BUY',
    }
    data_sell = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'TinkoffNew',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'SHIB',
    'fiat': 'RUB',
    'tradeType': 'SELL',
    }
    req = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data)
    req_sell = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data_sell)
#    print(req.text)

    book = load_workbook('binance.xlsx')
    sheet = book.active  

    """ПОКУПКА USDT ТИНЬКОФФ"""
    with open('data.json', 'w') as outfile:
        json.dump(req.json(), outfile)
    
    with open('data.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"buy: {pr['price']} rub")
        sheet['G3'] = pr['price']
        
    """ПРОДАЖА USDT ТИНЬКОФФ"""
    with open('data_sell.json', 'w') as outfile:
        json.dump(req_sell.json(), outfile)

    with open('data_sell.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"sell: {pr['price']} rub")
        sheet['M3'] = pr['price']


    book.save("binance.xlsx")

########################################################################################################
    """RosBankNew BNB"""
    data = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'RosBankNew',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'SHIB',
    'fiat': 'RUB',
    'tradeType': 'BUY',
    }
    data_sell = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'RosBankNew',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'SHIB',
    'fiat': 'RUB',
    'tradeType': 'SELL',
    }
    req = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data)
    req_sell = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data_sell)
#    print(req.text)
    book = load_workbook('binance.xlsx')
    sheet = book.active  


    """ПОКУПКА BUSD ТИНЬКОФФ"""
    with open('data.json', 'w') as outfile:
        json.dump(req.json(), outfile)
    
    with open('data.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"buy: {pr['price']} rub")
        sheet['G4'] = pr['price']

    """ПРОДАЖА USDT ТИНЬКОФФ"""
    with open('data_sell.json', 'w') as outfile:
        json.dump(req_sell.json(), outfile)

    with open('data_sell.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"sell: {pr['price']} rub")
        sheet['M4'] = pr['price']
    book.save("binance.xlsx")
########################################################################################################
    """QIWI BUSD"""
    data = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'QIWI',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'SHIB',
    'fiat': 'RUB',
    'tradeType': 'BUY',
    }
    data_sell = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'QIWI',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'SHIB',
    'fiat': 'RUB',
    'tradeType': 'SELL',
    }
    req = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data)
    req_sell = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data_sell)
#    print(req.text)
    book = load_workbook('binance.xlsx')
    sheet = book.active  


    """ПОКУПКА ETH ТИНЬКОФФ"""
    with open('data.json', 'w') as outfile:
        json.dump(req.json(), outfile)
    
    with open('data.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"buy: {pr['price']} rub")
        sheet['G5'] = pr['price']

    """ПРОДАЖА USDT ТИНЬКОФФ"""
    with open('data_sell.json', 'w') as outfile:
        json.dump(req_sell.json(), outfile)

    with open('data_sell.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"sell: {pr['price']} rub")
        sheet['M5'] = pr['price']

    book.save("binance.xlsx")

########################################################################################################
    """YANDEX MONEY BUSD"""
    data = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'YandexMoneyNew',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'SHIB',
    'fiat': 'RUB',
    'tradeType': 'BUY',
    }
    data_sell = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'YandexMoneyNew',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'SHIB',
    'fiat': 'RUB',
    'tradeType': 'SELL',
    }
    req = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data)
    req_sell = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data_sell)
#    print(req.text)
    book = load_workbook('binance.xlsx')
    sheet = book.active  


    """ПОКУПКА ETH ТИНЬКОФФ"""
    with open('data.json', 'w') as outfile:
        json.dump(req.json(), outfile)
    
    with open('data.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"buy: {pr['price']} rub")
        sheet['G6'] = pr['price']

    """ПРОДАЖА ETH ТИНЬКОФФ"""
    with open('data_sell.json', 'w') as outfile:
        json.dump(req_sell.json(), outfile)

    with open('data_sell.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"sell: {pr['price']} rub")
        sheet['M6'] = pr['price']

    book.save("binance.xlsx")
########################################################################################################
    """RAIFFAIZEN BUSD"""
    data = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'RaiffeisenBank',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'SHIB',
    'fiat': 'RUB',
    'tradeType': 'BUY',
    }
    data_sell = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'RaiffeisenBank',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'SHIB',
    'fiat': 'RUB',
    'tradeType': 'SELL',
    }
    req = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data)
    req_sell = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data_sell)
#    print(req.text)
    book = load_workbook('binance.xlsx')
    sheet = book.active  


    """ПОКУПКА ETH ТИНЬКОФФ"""
    with open('data.json', 'w') as outfile:
        json.dump(req.json(), outfile)
    
    with open('data.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"buy: {pr['price']} rub")
        sheet['G7'] = pr['price']

    """ПРОДАЖА ETH ТИНЬКОФФ"""
    with open('data_sell.json', 'w') as outfile:
        json.dump(req_sell.json(), outfile)

    with open('data_sell.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"sell: {pr['price']} rub")
        sheet['M7'] = pr['price']

    book.save("binance.xlsx")
########################################################################################################
    """ПОЧТА БАНК BUSD"""
    data = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'PostBankNew',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'SHIB',
    'fiat': 'RUB',
    'tradeType': 'BUY',
    }
    data_sell = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'PostBankNew',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'SHIB',
    'fiat': 'RUB',
    'tradeType': 'SELL',
    }
    req = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data)
    req_sell = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data_sell)
#    print(req.text)
    book = load_workbook('binance.xlsx')
    sheet = book.active  


    """ПОКУПКА ETH ТИНЬКОФФ"""
    with open('data.json', 'w') as outfile:
        json.dump(req.json(), outfile)
    
    with open('data.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"buy: {pr['price']} rub")
        sheet['G8'] = pr['price']

    """ПРОДАЖА USDT ТИНЬКОФФ"""
    with open('data_sell.json', 'w') as outfile:
        json.dump(req_sell.json(), outfile)

    with open('data_sell.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"sell: {pr['price']} rub")
        sheet['M8'] = pr['price']

    book.save("binance.xlsx")

########################################################################################################
    """МТС БАНК BUSD"""
    data = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'MTSBank',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'SHIB',
    'fiat': 'RUB',
    'tradeType': 'BUY',
    }
    data_sell = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'MTSBank',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'SHIB',
    'fiat': 'RUB',
    'tradeType': 'SELL',
    }
    req = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data)
    req_sell = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data_sell)
#    print(req.text)
    book = load_workbook('binance.xlsx')
    sheet = book.active  


    """ПОКУПКА ETH ТИНЬКОФФ"""
    with open('data.json', 'w') as outfile:
        json.dump(req.json(), outfile)
    
    with open('data.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"buy: {pr['price']} rub")
        sheet['G9'] = pr['price']

    """ПРОДАЖА USDT ТИНЬКОФФ"""
    with open('data_sell.json', 'w') as outfile:
        json.dump(req_sell.json(), outfile)

    with open('data_sell.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"sell: {pr['price']} rub")
        sheet['M9'] = pr['price']

    book.save("binance.xlsx")


########################################################################################################
    """ХОУМ-КРЕДИТ БАНК BUSD"""
    data = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'HomeCreditBank',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'SHIB',
    'fiat': 'RUB',
    'tradeType': 'BUY',
    }
    data_sell = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'HomeCreditBank',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'SHIB',
    'fiat': 'RUB',
    'tradeType': 'SELL',
    }
    req = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data)
    req_sell = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data_sell)
#    print(req.text)
    book = load_workbook('binance.xlsx')
    sheet = book.active  


    """ПОКУПКА ETH ТИНЬКОФФ"""
    with open('data.json', 'w') as outfile:
        json.dump(req.json(), outfile)
    
    with open('data.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"buy: {pr['price']} rub")
        sheet['G10'] = pr['price']

    """ПРОДАЖА USDT ТИНЬКОФФ"""
    with open('data_sell.json', 'w') as outfile:
        json.dump(req_sell.json(), outfile)

    with open('data_sell.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"sell: {pr['price']} rub")
        sheet['M10'] = pr['price']

    book.save("binance.xlsx")

########################################################################################################
    """ФИАТНЫЙ БАЛАНС БАНК BUSD"""
    data = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'RUBfiatbalance',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'SHIB',
    'fiat': 'RUB',
    'tradeType': 'BUY',
    }
    data_sell = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'RUBfiatbalance',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'SHIB',
    'fiat': 'RUB',
    'tradeType': 'SELL',
    }
    req = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data)
    req_sell = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data_sell)
#    print(req.text)
    book = load_workbook('binance.xlsx')
    sheet = book.active  


    """ПОКУПКА ETH ТИНЬКОФФ"""
    with open('data.json', 'w') as outfile:
        json.dump(req.json(), outfile)
    
    with open('data.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"buy: {pr['price']} rub")
        sheet['G12'] = pr['price']

    """ПРОДАЖА ETH ТИНЬКОФФ"""
    with open('data_sell.json', 'w') as outfile:
        json.dump(req_sell.json(), outfile)

    with open('data_sell.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"sell: {pr['price']} rub")
        sheet['M12'] = pr['price']

    book.save("binance.xlsx")
########################################################################################################
    """PAYER BUSD"""
    data = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'Payeer',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'SHIB',
    'fiat': 'RUB',
    'tradeType': 'BUY',
    }
    data_sell = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'Payeer',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'SHIB',
    'fiat': 'RUB',
    'tradeType': 'SELL',
    }
    req = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data)
    req_sell = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data_sell)
#    print(req.text)
    book = load_workbook('binance.xlsx')
    sheet = book.active  


    """ПОКУПКА ETH ТИНЬКОФФ"""
    with open('data.json', 'w') as outfile:
        json.dump(req.json(), outfile)
    
    with open('data.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"buy: {pr['price']} rub")
        sheet['G13'] = pr['price']

    """ПРОДАЖА ETH ТИНЬКОФФ"""
    with open('data_sell.json', 'w') as outfile:
        json.dump(req_sell.json(), outfile)

    with open('data_sell.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"sell: {pr['price']} rub")
        sheet['M13'] = pr['price']

    book.save("binance.xlsx")
########################################################################################################
    """ADVCASH BUSD"""
    data = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'Advcash',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'SHIB',
    'fiat': 'RUB',
    'tradeType': 'BUY',
    }
    data_sell = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'Advcash',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'SHIB',
    'fiat': 'RUB',
    'tradeType': 'SELL',
    }
    req = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data)
    req_sell = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data_sell)
#    print(req.text)
    book = load_workbook('binance.xlsx')
    sheet = book.active  


    """ПОКУПКА BUSD ТИНЬКОФФ"""
    with open('data.json', 'w') as outfile:
        json.dump(req.json(), outfile)
    
    with open('data.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"buy: {pr['price']} rub")
        sheet['G14'] = pr['price']

    """ПРОДАЖА BUSD ТИНЬКОФФ"""
    with open('data_sell.json', 'w') as outfile:
        json.dump(req_sell.json(), outfile)

    with open('data_sell.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"sell: {pr['price']} rub")
        sheet['M14'] = pr['price']

    book.save("binance.xlsx")
    book.close()


