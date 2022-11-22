import requests
import openpyxl
from openpyxl import load_workbook
import json
def parse():
    headers = {
        "Accept": "*/*",
        "Accept-Encoding": "gzip, deflate, br",
        "Accept-Language": "en-GB,en-US;q=0.9,en;q=0.8",
        "Cache-Control": "no-cache",
        "Connection": "keep-alive",
        "Content-Length": "123",
        "content-type": "application/json",
        "Host": "p2p.binance.com",
        "Origin": "https://p2p.binance.com",
        "Pragma": "no-cache",
        "TE": "Trailers",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:88.0) Gecko/20100101 Firefox/88.0"
    }

    data = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'TinkoffNew',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'USDT',
    'fiat': 'RUB',
    'tradeType': 'BUY',
    }
    data_sell = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'TinkoffNew',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'USDT',
    'fiat': 'RUB',
    'tradeType': 'SELL',
    }
    req = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data)
    req_sell = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data_sell)
#    print(req.text)
    book = load_workbook('binance.xlsx')
    sheet = book.active  


    """ПОКУПКА USDT ТИНЬКОФФ"""
    with open('data.json', 'w') as outfile:
        json.dump(req.json(), outfile)
    
    with open('data.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"buy: {pr['price']} rub")
        sheet['B3'] = float(pr['price'])
    """ПРОДАЖА USDT ТИНЬКОФФ"""
    with open('data_sell.json', 'w') as outfile:
        json.dump(req_sell.json(), outfile)

    with open('data_sell.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"sell: {pr['price']} rub")
        sheet['H3'] = float(pr['price'])

    book.save("binance.xlsx")

########################################################################################################
    """RosBankNew USDT"""
    data = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'RosBankNew',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'USDT',
    'fiat': 'RUB',
    'tradeType': 'BUY',
    }
    data_sell = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'RosBankNew',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'USDT',
    'fiat': 'RUB',
    'tradeType': 'SELL',
    }
    req = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data)
    req_sell = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data_sell)
#    print(req.text)
    book = load_workbook('binance.xlsx')
    sheet = book.active  


    """ПОКУПКА USDT ТИНЬКОФФ"""
    with open('data.json', 'w') as outfile:
        json.dump(req.json(), outfile)
    
    with open('data.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"buy: {pr['price']} rub")
        sheet['B4'] = float(pr['price'])

    """ПРОДАЖА USDT ТИНЬКОФФ"""
    with open('data_sell.json', 'w') as outfile:
        json.dump(req_sell.json(), outfile)

    with open('data_sell.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"sell: {pr['price']} rub")
        sheet['H4'] = float(pr['price'])

    book.save("binance.xlsx")
########################################################################################################
    """QIWI USDT"""
    data = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'QIWI',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'USDT',
    'fiat': 'RUB',
    'tradeType': 'BUY',
    }
    data_sell = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'QIWI',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'USDT',
    'fiat': 'RUB',
    'tradeType': 'SELL',
    }
    req = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data)
    req_sell = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data_sell)
#    print(req.text)
    book = load_workbook('binance.xlsx')
    sheet = book.active  


    """ПОКУПКА USDT ТИНЬКОФФ"""
    with open('data.json', 'w') as outfile:
        json.dump(req.json(), outfile)
    
    with open('data.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"buy: {pr['price']} rub")
        sheet['B5'] =float(pr['price'])

    """ПРОДАЖА USDT ТИНЬКОФФ"""
    with open('data_sell.json', 'w') as outfile:
        json.dump(req_sell.json(), outfile)

    with open('data_sell.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"sell: {pr['price']} rub")
        sheet['H5'] = float(pr['price'])

    book.save("binance.xlsx")

########################################################################################################
    """YANDEX MONEY USDT"""
    data = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'YandexMoneyNew',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'USDT',
    'fiat': 'RUB',
    'tradeType': 'BUY',
    }
    data_sell = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'YandexMoneyNew',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'USDT',
    'fiat': 'RUB',
    'tradeType': 'SELL',
    }
    req = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data)
    req_sell = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data_sell)
#    print(req.text)
    book = load_workbook('binance.xlsx')
    sheet = book.active  


    """ПОКУПКА USDT ТИНЬКОФФ"""
    with open('data.json', 'w') as outfile:
        json.dump(req.json(), outfile)
    
    with open('data.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"buy: {pr['price']} rub")
        sheet['B6'] = float(pr['price'])

    """ПРОДАЖА USDT ТИНЬКОФФ"""
    with open('data_sell.json', 'w') as outfile:
        json.dump(req_sell.json(), outfile)

    with open('data_sell.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"sell: {pr['price']} rub")
        sheet['H6'] = float(pr['price'])

    book.save("binance.xlsx")
########################################################################################################
    """RAIFFAIZEN USDT"""
    data = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'RaiffeisenBank',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'USDT',
    'fiat': 'RUB',
    'tradeType': 'BUY',
    }
    data_sell = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'RaiffeisenBank',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'USDT',
    'fiat': 'RUB',
    'tradeType': 'SELL',
    }
    req = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data)
    req_sell = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data_sell)
#    print(req.text)
    book = load_workbook('binance.xlsx')
    sheet = book.active  


    """ПОКУПКА USDT ТИНЬКОФФ"""
    with open('data.json', 'w') as outfile:
        json.dump(req.json(), outfile)
    
    with open('data.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"buy: {pr['price']} rub")
        sheet['B7'] = float(pr['price'])

    """ПРОДАЖА USDT ТИНЬКОФФ"""
    with open('data_sell.json', 'w') as outfile:
        json.dump(req_sell.json(), outfile)

    with open('data_sell.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"sell: {pr['price']} rub")
        sheet['H7'] = float(pr['price'])

    book.save("binance.xlsx")
########################################################################################################
    """ПОЧТА БАНК USDT"""
    data = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'PostBankNew',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'USDT',
    'fiat': 'RUB',
    'tradeType': 'BUY',
    }
    data_sell = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'PostBankNew',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'USDT',
    'fiat': 'RUB',
    'tradeType': 'SELL',
    }
    req = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data)
    req_sell = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data_sell)
#    print(req.text)
    book = load_workbook('binance.xlsx')
    sheet = book.active  


    """ПОКУПКА USDT ТИНЬКОФФ"""
    with open('data.json', 'w') as outfile:
        json.dump(req.json(), outfile)
    
    with open('data.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"buy: {pr['price']} rub")
        sheet['B8'] = float(pr['price'])

    """ПРОДАЖА USDT ТИНЬКОФФ"""
    with open('data_sell.json', 'w') as outfile:
        json.dump(req_sell.json(), outfile)

    with open('data_sell.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"sell: {pr['price']} rub")
        sheet['H8'] = float(pr['price'])

    book.save("binance.xlsx")

########################################################################################################
    """МТС БАНК USDT"""
    data = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'MTSBank',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'USDT',
    'fiat': 'RUB',
    'tradeType': 'BUY',
    }
    data_sell = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'MTSBank',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'USDT',
    'fiat': 'RUB',
    'tradeType': 'SELL',
    }
    req = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data)
    req_sell = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data_sell)
#    print(req.text)
    book = load_workbook('binance.xlsx')
    sheet = book.active  


    """ПОКУПКА USDT ТИНЬКОФФ"""
    with open('data.json', 'w') as outfile:
        json.dump(req.json(), outfile)
    
    with open('data.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"buy: {pr['price']} rub")
        sheet['B9'] = float(pr['price'])

    """ПРОДАЖА USDT ТИНЬКОФФ"""
    with open('data_sell.json', 'w') as outfile:
        json.dump(req_sell.json(), outfile)

    with open('data_sell.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"sell: {pr['price']} rub")
        sheet['H9'] = float(pr['price'])

    book.save("binance.xlsx")


########################################################################################################
    """ХОУМ-КРЕДИТ БАНК USDT"""
    data = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'HomeCreditBank',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'USDT',
    'fiat': 'RUB',
    'tradeType': 'BUY',
    }
    data_sell = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'HomeCreditBank',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'USDT',
    'fiat': 'RUB',
    'tradeType': 'SELL',
    }
    req = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data)
    req_sell = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data_sell)
#    print(req.text)
    book = load_workbook('binance.xlsx')
    sheet = book.active  


    """ПОКУПКА USDT ТИНЬКОФФ"""
    with open('data.json', 'w') as outfile:
        json.dump(req.json(), outfile)
    
    with open('data.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"buy: {pr['price']} rub")
        sheet['B10'] = float(pr['price'])

    """ПРОДАЖА USDT ТИНЬКОФФ"""
    with open('data_sell.json', 'w') as outfile:
        json.dump(req_sell.json(), outfile)

    with open('data_sell.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"sell: {pr['price']} rub")
        sheet['H10'] = float(pr['price'])

    book.save("binance.xlsx")

########################################################################################################
    """ФИАТНЫЙ БАЛАНС БАНК USDT"""
    data = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'RUBfiatbalance',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'USDT',
    'fiat': 'RUB',
    'tradeType': 'BUY',
    }
    data_sell = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'RUBfiatbalance',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'USDT',
    'fiat': 'RUB',
    'tradeType': 'SELL',
    }
    req = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data)
    req_sell = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data_sell)
#    print(req.text)
    book = load_workbook('binance.xlsx')
    sheet = book.active  


    """ПОКУПКА USDT ТИНЬКОФФ"""
    with open('data.json', 'w') as outfile:
        json.dump(req.json(), outfile)
    
    with open('data.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"buy: {pr['price']} rub")
        sheet['B12'] = float(pr['price'])

    """ПРОДАЖА USDT ТИНЬКОФФ"""
    with open('data_sell.json', 'w') as outfile:
        json.dump(req_sell.json(), outfile)

    with open('data_sell.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"sell: {pr['price']} rub")
        sheet['H12'] = float(pr['price'])
    book.save("binance.xlsx")
########################################################################################################
    """PAYER USDT"""
    data = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'Payeer',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'USDT',
    'fiat': 'RUB',
    'tradeType': 'BUY',
    }
    data_sell = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'Payeer',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'USDT',
    'fiat': 'RUB',
    'tradeType': 'SELL',
    }
    req = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data)
    req_sell = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data_sell)
#    print(req.text)
    book = load_workbook('binance.xlsx')
    sheet = book.active  


    """ПОКУПКА USDT ТИНЬКОФФ"""
    with open('data.json', 'w') as outfile:
        json.dump(req.json(), outfile)
    
    with open('data.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"buy: {pr['price']} rub")
        sheet['B13'] = float(pr['price'])

    """ПРОДАЖА USDT ТИНЬКОФФ"""
    with open('data_sell.json', 'w') as outfile:
        json.dump(req_sell.json(), outfile)

    with open('data_sell.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"sell: {pr['price']} rub")
        sheet['H13'] = float(pr['price'])

    book.save("binance.xlsx")
########################################################################################################
    """ADVCASH USDT"""
    data = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'Advcash',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'USDT',
    'fiat': 'RUB',
    'tradeType': 'BUY',
    }
    data_sell = {
    'page': 1,
    'rows': 1,
    'payTypes': [
        'Advcash',
    ],
    'countries': [],
    'publisherType': None,
    'asset': 'USDT',
    'fiat': 'RUB',
    'tradeType': 'SELL',
    }
    req = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data)
    req_sell = requests.post('https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search', headers=headers,json=data_sell)
#    print(req.text)
    book = load_workbook('binance.xlsx')
    sheet = book.active  


    """ПОКУПКА USDT ТИНЬКОФФ"""
    with open('data.json', 'w') as outfile:
        json.dump(req.json(), outfile)
    
    with open('data.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"buy: {pr['price']} rub")
        sheet['B14'] = float(pr['price'])

    """ПРОДАЖА USDT ТИНЬКОФФ"""
    with open('data_sell.json', 'w') as outfile:
        json.dump(req_sell.json(), outfile)

    with open('data_sell.json') as file:
        jsData = json.load(file)

    for price in jsData['data']:
        pr = price['adv']
        print(f"sell: {pr['price']} rub")
        sheet['H14'] = float(pr['price'])

    book.save("binance.xlsx")
    book.close()
